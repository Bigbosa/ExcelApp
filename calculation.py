import tkinter as tk
from tkinter import filedialog
import openpyxl
from customtkinter import CTkButton, CTkFrame, CTkLabel, CTkEntry  # Adjust imports based on your customtkinter library

class ExcelCalculatorApp(tk.Toplevel):
    def __init__(self, master=None, *args, **kwargs):
        tk.Toplevel.__init__(self, master, *args, **kwargs)
        self.title("Excel Calculator")
        self.geometry("800x380")
        self.configure(bg="darkgreen")
        self.iconbitmap("ExcelApp_logo.ico")
        self.resizable(False, False)

        # Create buttons for SUM, MIN, AVERAGE, MAX, COUNT
        sum_button = CTkButton(self, text="Calculate SUM", command=lambda: self.upload_file("SUM"), hover_color="darkgreen",
                               bg_color="white", fg_color="transparent", text_color="black", border_color="darkgreen", border_width=3)
        sum_button.pack(side=tk.TOP, padx=10, pady=10)

        min_button = CTkButton(self, text="Calculate MIN", command=lambda: self.upload_file("MIN"),hover_color="darkgreen",
                               bg_color="white", fg_color="transparent", text_color="black", border_color="darkgreen", border_width=3)
        min_button.pack(side=tk.TOP, padx=10, pady=10)

        average_button = CTkButton(self, text="Calculate AVERAGE", command=lambda: self.upload_file("AVERAGE"),
                                   hover_color="darkgreen",
                                   bg_color="white", fg_color="transparent", text_color="black",
                                   border_color="darkgreen", border_width=3)
        average_button.pack(side=tk.TOP, padx=10, pady=10)

        max_button = CTkButton(self, text="Calculate MAX", command=lambda: self.upload_file("MAX"), hover_color="darkgreen",
                               bg_color="white", fg_color="transparent", text_color="black", border_color="darkgreen", border_width=3)
        max_button.pack(side=tk.TOP, padx=10, pady=10)

        count_button = CTkButton(self, text="Calculate COUNT", command=lambda: self.upload_file("COUNT"),
                                 hover_color="darkgreen",
                                 bg_color="white", fg_color="transparent", text_color="black", border_color="darkgreen",
                                 border_width=3)
        count_button.pack(side=tk.TOP, padx=10, pady=10)

        # Create a custom frame with a big textbox
        text_frame = CTkFrame(self, bg_color="darkgreen")
        text_frame.pack(side=tk.BOTTOM, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Add a label above the big textbox
        notes_label = CTkLabel(text_frame, text="Notes",  text_color="black", font=("Arial", 20, "bold"))
        notes_label.pack(pady=5)

        # Add a big textbox to the frame
        big_textbox = tk.Text(text_frame, wrap=tk.WORD, height=10, width=50, bg="white", fg="black", font=("Arial", 12))
        big_textbox.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Initialize the file_path variable and result_label
        self.file_path = None
        self.result_label = CTkLabel(self, text="", fg_color="darkgreen", text_color="white")
        self.result_label.pack(side=tk.TOP, padx=10, pady=10)

    def calculate_and_save(self, selected_column, result_row, calculation_type):
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            # Find the last row with data
            last_row = sheet.max_row

            # Validate the selected column
            if not selected_column.isalpha():
                self.result_label.config(text="Error: Please enter a valid column letter (e.g., 'D').")
                return

            # Calculate the formula based on the selected type
            if calculation_type == "SUM":
                formula = f"=SUM({selected_column}2:{selected_column}{last_row})"
            elif calculation_type == "MIN":
                formula = f"=MIN({selected_column}2:{selected_column}{last_row})"
            elif calculation_type == "AVERAGE":
                formula = f"=AVERAGE({selected_column}2:{selected_column}{last_row})"
            elif calculation_type == "MAX":
                formula = f"=MAX({selected_column}2:{selected_column}{last_row})"
            elif calculation_type == "COUNT":
                formula = f"=COUNT({selected_column}2:{selected_column}{last_row})"
            else:
                self.result_label.config(text="Error: Invalid calculation type.")
                return

            # Apply the formula to the specified row and column
            for col_num, cell_value in enumerate(
                    sheet.iter_rows(min_row=result_row, max_row=result_row, values_only=True), start=1):
                sheet.cell(row=result_row, column=col_num, value=formula)

            # Save the updated workbook
            output_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                       filetypes=[("Excel files", "*.xlsx;*.xls")])
            wb.save(output_path)

            self.result_label.config(text=f"{calculation_type} calculated and saved to {output_path}")

        except Exception as e:
            self.result_label.config(text=f"Error: {e}")

    def select_column(self, calculation_type):
        # Create a new window for column selection
        column_window = tk.Toplevel(self)
        column_window.title("Select Column")
        column_window.configure(bg="black")

        # Label and entry for column selection
        column_label = CTkLabel(column_window, text="Enter Column Letter:", bg="darkgreen", fg="white")
        column_label.pack()

        column_entry = CTkEntry(column_window, bg="darkgreen", fg="white")
        column_entry.pack()

        # Button to confirm column selection and prompt for row
        confirm_button = CTkButton(column_window, text="Confirm", command=lambda: self.select_row(column_entry.get(), calculation_type), bg="red", fg="black")
        confirm_button.pack()

    def select_row(self, selected_column, calculation_type):
        # Create a new window for row selection
        row_window = tk.Toplevel(self)
        row_window.title("Enter Row Number")
        row_window.configure(bg="darkgreen")

        # Label and entry for row selection
        row_label = CTkLabel(row_window, text="Enter Row Number:", bg="darkgreen", fg="white")
        row_label.pack()

        row_entry = CTkEntry(row_window, bg="darkgreen", fg="white")
        row_entry.pack()

        # Button to confirm row selection and perform calculation
        confirm_button = CTkButton(row_window, text="Confirm", command=lambda: self.on_confirm(selected_column, int(row_entry.get()), calculation_type, row_window), bg="red", fg="black")
        confirm_button.pack()

    def upload_file(self, calculation_type):
        # Get the selected Excel file
        self.file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")])

        if self.file_path:
            self.select_column(calculation_type)

    def on_confirm(self, selected_column, result_row, calculation_type, window):
        self.calculate_and_save(selected_column, result_row, calculation_type)
        window.destroy()

if __name__ == "__main__":
    excel_calculator_app = ExcelCalculatorApp()
    excel_calculator_app.mainloop()
