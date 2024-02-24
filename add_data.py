from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from customtkinter import CTkButton, CTkFrame, CTkLabel
import tkinter as tk

class AddDataFrame(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)

        self.title("Add Data")
        self.geometry("800x380")
        self.configure(bg="darkgreen")
        self.iconbitmap("ExcelApp_logo.ico")
        self.file_path = None

        self.resizable(False, False)

        # Create a frame to hold widgets
        self.frame = tk.Frame(self, padx=10, pady=10, bg="darkgreen")
        self.frame.pack()

        # Create a button to upload file
        self.upload_button = CTkButton(self.frame, text="Upload Excel File", command=self.upload_file,
                                       hover_color="darkgreen",
                                       fg_color="transparent", text_color="white", border_color="darkgreen",
                                       border_width=2)
        self.upload_button.grid(row=1, column=0, pady=5)

        # Create an entry for data input
        self.data_entry = tk.Entry(self.frame, width=30)
        self.data_entry.grid(row=1, column=1, pady=5)

        # Create a button to add data
        self.add_data_button = CTkButton(self.frame, text="Add Data", command=self.add_data, hover_color="darkgreen",
                                         fg_color="transparent", text_color="white", border_color="darkgreen",
                                         border_width=2)
        self.add_data_button.grid(row=1, column=2, pady=5)

        # Create a custom frame with a big textbox
        text_frame = CTkFrame(self, bg_color="darkgreen")
        text_frame.pack(side=tk.BOTTOM, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Add a label above the big textbox
        notes_label = CTkLabel(text_frame, text="Notes", text_color="black", font=("Arial", 20, "bold"))
        notes_label.pack(pady=5)

        # Add a big textbox to the frame
        self.big_textbox = tk.Text(text_frame, wrap=tk.WORD, height=10, width=50, bg="white", fg="black",
                                   font=("Arial", 12))
        self.big_textbox.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.BOTH, expand=True)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.file_path = file_path


    def add_data(self):
        if self.file_path:
            data_to_add = self.data_entry.get()

            # Load existing workbook or create a new one
            try:
                workbook = load_workbook(self.file_path)
            except FileNotFoundError:
                workbook = Workbook()

            # Select the active sheet (the first sheet in the workbook)
            sheet = workbook.active

            # Add headers if the sheet is empty
            if sheet.max_row == 1:
                headers = ["Additional Data"]
                for col_num, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    sheet[f"{col_letter}1"] = header
                    sheet[f"{col_letter}1"].font = Font(bold=True)

            # Append data to the sheet
            next_row = sheet.max_row + 1
            sheet[f"A{next_row}"] = data_to_add

            # Save the workbook
            workbook.save(self.file_path)

if __name__ == "__main__":
    add_data_frame = AddDataFrame()
    add_data_frame.mainloop()
